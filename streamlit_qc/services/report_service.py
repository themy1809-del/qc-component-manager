# -*- coding: utf-8 -*-
"""
Service: báo cáo + xuất Excel + chart tiến độ.

Tương đương `_export_report` trong Tkinter (dòng 1409-1429) + bổ sung chart.
"""
from __future__ import annotations

import datetime as dt
import io
import json
from dataclasses import dataclass, field

import pandas as pd

from streamlit_qc.core.db import DB


@dataclass
class WeeklyProgress:
    """Tiến độ theo tuần."""
    week_start: str            # ISO date (Monday)
    week_label: str            # vd "T20 (12-18/05)"
    inspections: int           # số inspection trong tuần
    components_inspected: int  # số cấu kiện unique được inspect
    cumulative: int = 0        # số inspection cộng dồn


@dataclass
class ReportData:
    """Tất cả số liệu cho page Báo cáo."""
    # Số liệu tổng quan trong range
    total_components: int = 0
    total_inspections: int = 0
    accepted: int = 0
    passed: int = 0
    failed: int = 0

    # Theo tuần
    weekly: list[WeeklyProgress] = field(default_factory=list)

    # Theo NT type
    by_type: dict[str, int] = field(default_factory=dict)

    # Top xưởng theo % hoàn thành
    workshop_progress: list[dict] = field(default_factory=list)

    # Theo người kiểm tra
    by_inspector: list[dict] = field(default_factory=list)


def get_inspection_date_range(db: DB, pid: int) -> tuple[dt.date | None, dt.date | None]:
    """Trả về (min_date, max_date) của tất cả inspection trong dự án."""
    row = db.conn.execute(
        "SELECT MIN(inspection_date) mn, MAX(inspection_date) mx "
        "FROM inspections WHERE project_id=? AND inspection_date IS NOT NULL",
        (pid,),
    ).fetchone()
    if not row or not row["mn"]:
        return None, None
    try:
        mn = dt.date.fromisoformat(row["mn"][:10])
        mx = dt.date.fromisoformat(row["mx"][:10])
        return mn, mx
    except ValueError:
        return None, None


def compute_report(
    db: DB,
    pid: int,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> ReportData:
    """
    Tính số liệu báo cáo trong khoảng date.

    Args:
        date_from, date_to: nếu None = không lọc.
    """
    data = ReportData()

    # ---- Components total (luôn lấy hết, không lọc theo ngày) ----
    data.total_components = db.conn.execute(
        "SELECT COUNT(*) c FROM components WHERE project_id=?", (pid,)
    ).fetchone()["c"]

    # ---- Inspection query với date filter ----
    sql = (
        "SELECT i.*, c.code comp_code, c.data_json comp_data "
        "FROM inspections i JOIN components c ON c.id=i.component_id "
        "WHERE i.project_id=?"
    )
    params: list = [pid]
    if date_from:
        sql += " AND i.inspection_date >= ?"
        params.append(date_from.isoformat())
    if date_to:
        sql += " AND i.inspection_date <= ?"
        params.append(date_to.isoformat())
    sql += " ORDER BY i.inspection_date ASC, i.id ASC"

    rows = db.conn.execute(sql, params).fetchall()
    data.total_inspections = len(rows)

    # ---- Components: status counts (không theo range) ----
    for r in db.conn.execute(
        "SELECT status, COUNT(*) c FROM components WHERE project_id=? GROUP BY status",
        (pid,),
    ).fetchall():
        if r["status"] == "ACCEPTED":
            data.accepted = r["c"]
        elif r["status"] == "PASSED":
            data.passed = r["c"]
        elif r["status"] == "FAILED":
            data.failed = r["c"]

    # ---- Aggregate theo tuần + theo NT type + theo inspector ----
    weekly_map: dict[dt.date, dict] = {}
    type_map: dict[str, int] = {}
    inspector_map: dict[str, dict] = {}
    workshop_inspected: dict[str, set[int]] = {}

    for r in rows:
        # Parse date
        try:
            d = dt.date.fromisoformat(r["inspection_date"][:10])
        except (ValueError, TypeError):
            continue

        # Weekly
        # ISO week: Monday-based
        monday = d - dt.timedelta(days=d.weekday())
        if monday not in weekly_map:
            weekly_map[monday] = {"inspections": 0, "components": set()}
        weekly_map[monday]["inspections"] += 1
        weekly_map[monday]["components"].add(r["component_id"])

        # By type
        t = r["inspection_type"]
        type_map[t] = type_map.get(t, 0) + 1

        # By inspector
        ins = (r["inspector"] or "(không rõ)").strip()
        if ins not in inspector_map:
            inspector_map[ins] = {"inspections": 0, "components": set()}
        inspector_map[ins]["inspections"] += 1
        inspector_map[ins]["components"].add(r["component_id"])

        # Workshop inspected
        try:
            comp_data = json.loads(r["comp_data"])
            ws = str(comp_data.get("workshop") or "(không xưởng)")
        except (json.JSONDecodeError, TypeError):
            ws = "(không xưởng)"
        if ws not in workshop_inspected:
            workshop_inspected[ws] = set()
        workshop_inspected[ws].add(r["component_id"])

    # Sort weekly và tính cumulative
    cumulative = 0
    for monday in sorted(weekly_map.keys()):
        w = weekly_map[monday]
        cumulative += w["inspections"]
        sunday = monday + dt.timedelta(days=6)
        # Tuần ISO
        iso_week = monday.isocalendar().week
        data.weekly.append(WeeklyProgress(
            week_start=monday.isoformat(),
            week_label=f"T{iso_week} ({monday.day:02d}-{sunday.day:02d}/{monday.month:02d})",
            inspections=w["inspections"],
            components_inspected=len(w["components"]),
            cumulative=cumulative,
        ))

    data.by_type = dict(sorted(type_map.items(), key=lambda x: -x[1]))

    data.by_inspector = sorted(
        [
            {
                "inspector": ins,
                "inspections": v["inspections"],
                "components": len(v["components"]),
            }
            for ins, v in inspector_map.items()
        ],
        key=lambda x: -x["inspections"],
    )

    # Workshop progress: cần tổng cấu kiện mỗi xưởng + số đã inspect
    ws_total: dict[str, int] = {}
    for r in db.conn.execute(
        "SELECT data_json FROM components WHERE project_id=?", (pid,)
    ):
        try:
            d = json.loads(r["data_json"])
            ws = str(d.get("workshop") or "(không xưởng)")
            ws_total[ws] = ws_total.get(ws, 0) + 1
        except json.JSONDecodeError:
            pass

    rows_ws = []
    for ws, total in sorted(ws_total.items()):
        inspected = len(workshop_inspected.get(ws, set()))
        pct = round(inspected * 100 / total, 1) if total else 0.0
        rows_ws.append({
            "workshop": ws,
            "total": total,
            "inspected_in_range": inspected,
            "percent_in_range": pct,
        })
    data.workshop_progress = sorted(rows_ws, key=lambda x: -x["percent_in_range"])

    return data


def export_to_excel_pro(
    db: DB,
    pid: int,
    project_code: str,
    project_name: str = "",
    overdue_threshold: int = 7,
) -> bytes:
    """
    Xuất báo cáo Excel CHUYÊN NGHIỆP 6 sheet với format đẹp.

    Sheets:
    1. 📋 Tổng quan — KPI + summary
    2. 🔧 Cấu kiện — full list với status
    3. ⚠️ Overdue — cấu kiện Fit-up quá N ngày chưa Final
    4. ❌ FAIL — cấu kiện không đạt
    5. 👤 Inspector — hiệu suất inspector
    6. 📋 Inspections — raw inspection records

    Returns: bytes Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Color palette
    NAVY = "0F1E40"
    GOLD = "D4A744"
    SUCCESS = "0F766E"
    WARNING = "D97706"
    DANGER = "DC2626"
    LIGHT_BG = "F8FAFC"

    # Common styles
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color=NAVY)
    sub_font = Font(italic=True, size=9, color="64748B")
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ==========================================================
    # SHEET 1: TỔNG QUAN
    # ==========================================================
    ws = wb.create_sheet("📋 Tổng quan")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    ws["A1"] = "BÁO CÁO QC COMPONENT MANAGER"
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Dự án: {project_code} - {project_name}"
    ws["A2"].font = Font(bold=True, size=11, color="475569")
    ws.merge_cells("A2:B2")
    ws["A3"] = f"Ngày xuất: {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = sub_font
    ws.merge_cells("A3:B3")

    counts = db.count_status(pid)
    total = counts.get("TOTAL", 0)
    accepted = counts.get("ACCEPTED", 0)
    passed = counts.get("PASSED", 0)
    pending = counts.get("PENDING", 0)
    in_progress = counts.get("IN_PROGRESS", 0)
    failed = counts.get("FAILED", 0)
    done = accepted + passed
    pct = round(done * 100 / total, 1) if total else 0.0

    rows = [
        ("", ""),
        ("Chỉ số kiểm tra", "Số lượng"),
        ("Tổng cấu kiện", total),
        ("Chưa kiểm tra (PENDING)", pending),
        ("Đã Fit-up (IN_PROGRESS)", in_progress),
        ("Đạt - DIR+VIR+NDT (PASSED)", passed),
        ("Đã nghiệm thu (ACCEPTED)", accepted),
        ("Không đạt (FAILED)", failed),
        ("", ""),
        ("% Hoàn thành", f"{pct}%"),
    ]
    for i, (k, v) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if i == 6:  # header row
            ws.cell(row=i, column=1).fill = header_fill
            ws.cell(row=i, column=2).fill = header_fill
            ws.cell(row=i, column=1).font = header_font
            ws.cell(row=i, column=2).font = header_font

    # ==========================================================
    # SHEET 2: CẤU KIỆN
    # ==========================================================
    ws_c = wb.create_sheet("🔧 Cấu kiện")
    comp_rows = db.conn.execute(
        "SELECT id, code, status, data_json FROM components WHERE project_id=? ORDER BY code",
        (pid,),
    ).fetchall()
    comp_records = []
    for r in comp_rows:
        try:
            d = json.loads(r["data_json"])
        except json.JSONDecodeError:
            d = {}
        comp_records.append({
            "Mã cấu kiện": r["code"],
            "Tên / Bản vẽ": d.get("drawing") or d.get("manual_drawing") or d.get("member_no") or "",
            "Revision": d.get("rev_no") or "",
            "Xưởng": d.get("workshop") or "",
            "Vật liệu": d.get("material") or "",
            "Section": d.get("section") or "",
            "Length (mm)": d.get("length_mm") or "",
            "Weight (kg)": d.get("weight_kg") or "",
            "Trạng thái": r["status"],
        })
    df_comp = pd.DataFrame(comp_records)
    if not df_comp.empty:
        for c_idx, col_name in enumerate(df_comp.columns, start=1):
            cell = ws_c.cell(row=1, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r_idx, row_data in enumerate(df_comp.values.tolist(), start=2):
            for c_idx, val in enumerate(row_data, start=1):
                ws_c.cell(row=r_idx, column=c_idx, value=val).border = border
        # Auto width
        for col_idx, col_name in enumerate(df_comp.columns, start=1):
            ws_c.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(col_name)) + 2)
        ws_c.freeze_panes = "A2"

    # ==========================================================
    # SHEET 3: OVERDUE
    # ==========================================================
    ws_o = wb.create_sheet("⚠️ Overdue")
    from streamlit_qc.services.component_service import get_overdue_components
    overdue_list = get_overdue_components(db, pid, overdue_threshold)
    ws_o["A1"] = f"Cấu kiện đã Fit-up > {overdue_threshold} ngày chưa Final"
    ws_o["A1"].font = title_font
    ws_o.merge_cells("A1:F1")
    ws_o["A2"] = f"Tổng: {len(overdue_list)} cấu kiện"
    ws_o["A2"].font = Font(bold=True, color=WARNING)

    if overdue_list:
        df_overdue = pd.DataFrame(overdue_list)
        df_overdue = df_overdue.rename(columns={
            "code": "Mã cấu kiện", "name": "Bản vẽ",
            "workshop": "Xưởng", "fitup_date": "Ngày Fit-up",
            "days_overdue": "Ngày quá hạn", "status": "Trạng thái",
        })
        cols_show = ["Mã cấu kiện", "Bản vẽ", "Xưởng", "Ngày Fit-up", "Ngày quá hạn", "Trạng thái"]
        df_overdue = df_overdue[[c for c in cols_show if c in df_overdue.columns]]

        for c_idx, col_name in enumerate(df_overdue.columns, start=1):
            cell = ws_o.cell(row=4, column=c_idx, value=col_name)
            cell.fill = PatternFill(start_color=WARNING, end_color=WARNING, fill_type="solid")
            cell.font = header_font
        for r_idx, row_data in enumerate(df_overdue.values.tolist(), start=5):
            for c_idx, val in enumerate(row_data, start=1):
                ws_o.cell(row=r_idx, column=c_idx, value=val).border = border
        for col_idx, col_name in enumerate(df_overdue.columns, start=1):
            ws_o.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col_name)) + 2)

    # ==========================================================
    # SHEET 4: FAIL
    # ==========================================================
    ws_f = wb.create_sheet("❌ FAIL")
    fail_rows = db.conn.execute(
        """SELECT c.code, c.status, c.data_json,
                  i.inspection_type, i.inspection_date, i.inspector, i.report_no, i.note
           FROM inspections i
           JOIN components c ON c.id = i.component_id
           WHERE i.project_id = ? AND i.result = 'FAIL'
           ORDER BY i.inspection_date DESC""",
        (pid,),
    ).fetchall()
    ws_f["A1"] = "Cấu kiện KHÔNG ĐẠT (FAIL) - cần xử lý"
    ws_f["A1"].font = title_font
    ws_f.merge_cells("A1:G1")
    ws_f["A2"] = f"Tổng: {len(fail_rows)} inspection FAIL"
    ws_f["A2"].font = Font(bold=True, color=DANGER)

    if fail_rows:
        fail_data = []
        for r in fail_rows:
            try:
                d = json.loads(r["data_json"])
            except json.JSONDecodeError:
                d = {}
            fail_data.append({
                "Mã cấu kiện": r["code"],
                "Loại NT": r["inspection_type"],
                "Ngày KT": r["inspection_date"] or "",
                "Inspector": r["inspector"] or "",
                "Báo cáo": r["report_no"] or "",
                "Ghi chú": (r["note"] or "")[:100],
                "Xưởng": d.get("workshop") or "",
                "Trạng thái HT": r["status"],
            })
        df_fail = pd.DataFrame(fail_data)
        for c_idx, col_name in enumerate(df_fail.columns, start=1):
            cell = ws_f.cell(row=4, column=c_idx, value=col_name)
            cell.fill = PatternFill(start_color=DANGER, end_color=DANGER, fill_type="solid")
            cell.font = header_font
        for r_idx, row_data in enumerate(df_fail.values.tolist(), start=5):
            for c_idx, val in enumerate(row_data, start=1):
                ws_f.cell(row=r_idx, column=c_idx, value=val).border = border
        for col_idx, col_name in enumerate(df_fail.columns, start=1):
            ws_f.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col_name)) + 2)

    # ==========================================================
    # SHEET 5: INSPECTOR PERFORMANCE
    # ==========================================================
    ws_i = wb.create_sheet("👤 Inspector")
    ws_i["A1"] = "Hiệu suất Inspector (30 ngày qua)"
    ws_i["A1"].font = title_font
    ws_i.merge_cells("A1:F1")

    from streamlit_qc.services.dashboard_service import get_inspector_performance
    perf = get_inspector_performance(db, pid, days=30)
    if perf:
        df_perf = pd.DataFrame(perf)
        df_perf["fail_rate"] = (df_perf["n_fail"] * 100 / df_perf["total"]).round(1)
        df_perf = df_perf.rename(columns={
            "inspector": "Inspector", "total": "Tổng",
            "n_pass": "Pass", "n_fail": "Fail",
            "n_recheck": "Recheck", "fail_rate": "% Fail",
        })
        for c_idx, col_name in enumerate(df_perf.columns, start=1):
            cell = ws_i.cell(row=3, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        for r_idx, row_data in enumerate(df_perf.values.tolist(), start=4):
            for c_idx, val in enumerate(row_data, start=1):
                ws_i.cell(row=r_idx, column=c_idx, value=val).border = border
        for col_idx, col_name in enumerate(df_perf.columns, start=1):
            ws_i.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col_name)) + 2)

    # ==========================================================
    # SHEET 6: INSPECTIONS (raw)
    # ==========================================================
    ws_ins = wb.create_sheet("📋 Inspections")
    df_ins = pd.read_sql_query(
        "SELECT i.id, c.code 'Mã cấu kiện', i.inspection_type 'Loại', "
        "i.inspection_date 'Ngày', i.inspector 'Inspector', i.result 'Kết quả', "
        "i.report_no 'Báo cáo', i.rfi_no 'RFI', i.source_file 'Nguồn' "
        "FROM inspections i "
        "JOIN components c ON c.id = i.component_id "
        "WHERE i.project_id = ? "
        "ORDER BY i.id DESC LIMIT 50000",
        db.conn, params=[pid],
    )
    if not df_ins.empty:
        for c_idx, col_name in enumerate(df_ins.columns, start=1):
            cell = ws_ins.cell(row=1, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        for r_idx, row_data in enumerate(df_ins.values.tolist(), start=2):
            for c_idx, val in enumerate(row_data, start=1):
                ws_ins.cell(row=r_idx, column=c_idx, value=val)
        for col_idx, col_name in enumerate(df_ins.columns, start=1):
            ws_ins.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(col_name)) + 2)
        ws_ins.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    db.log("system", "EXPORT_REPORT_PRO", "project", pid,
           f"sheets=6, components={len(comp_records)}, overdue={len(overdue_list)}, fail={len(fail_rows)}")
    return buf.getvalue()


def export_to_excel(db: DB, pid: int, project_code: str) -> bytes:
    """
    Xuất báo cáo ra file Excel 3 sheet: Components / Inspections / Summary.

    Returns:
        Bytes của file Excel (để dùng với st.download_button).
    """
    # Sheet 1: Components (tất cả 24 trường + status + audit fields)
    comp_rows = db.conn.execute(
        "SELECT id, code, status, data_json FROM components WHERE project_id=?",
        (pid,),
    ).fetchall()
    comp_records = []
    for r in comp_rows:
        try:
            d = json.loads(r["data_json"])
        except json.JSONDecodeError:
            d = {}
        d["__id"] = r["id"]
        d["__code"] = r["code"]
        d["__status"] = r["status"]
        comp_records.append(d)
    df_comp = pd.DataFrame(comp_records)
    # Đưa __code, __status lên đầu
    if "__code" in df_comp.columns:
        cols = ["__code", "__status"] + [c for c in df_comp.columns if c not in ("__code", "__status", "__id")]
        df_comp = df_comp[cols]

    # Sheet 2: Inspections (raw)
    df_ins = pd.read_sql_query(
        "SELECT i.*, c.code component_code "
        "FROM inspections i JOIN components c ON c.id=i.component_id "
        "WHERE i.project_id=? "
        "ORDER BY i.inspection_date DESC, i.id DESC",
        db.conn,
        params=[pid],
    )

    # Sheet 3: Summary (status counts)
    status_counts = db.count_status(pid)
    df_sum = pd.DataFrame([{
        "Mã dự án": project_code,
        "Tổng cấu kiện": status_counts.get("TOTAL", 0),
        "Chưa KT (PENDING)": status_counts.get("PENDING", 0),
        "Đang KT (IN_PROGRESS)": status_counts.get("IN_PROGRESS", 0),
        "Đạt (PASSED)": status_counts.get("PASSED", 0),
        "Không đạt (FAILED)": status_counts.get("FAILED", 0),
        "Đã nghiệm thu (ACCEPTED)": status_counts.get("ACCEPTED", 0),
        "Tổng inspection records": len(df_ins),
        "Ngày xuất": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
    }])

    # Sheet 4: Inspection by type (thêm để báo cáo nghiệp vụ)
    by_type = pd.read_sql_query(
        "SELECT inspection_type 'Loại NT', COUNT(*) 'Số inspection' "
        "FROM inspections WHERE project_id=? "
        "GROUP BY inspection_type ORDER BY COUNT(*) DESC",
        db.conn,
        params=[pid],
    )

    # Write to bytes
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_sum.to_excel(w, sheet_name="Tổng quan", index=False)
        df_comp.to_excel(w, sheet_name="Cấu kiện", index=False)
        df_ins.to_excel(w, sheet_name="Inspections", index=False)
        by_type.to_excel(w, sheet_name="Theo loại NT", index=False)

    db.log("system", "EXPORT_REPORT", "project", pid,
           f"components={len(df_comp)}, inspections={len(df_ins)}")
    return buf.getvalue()
