# -*- coding: utf-8 -*-
"""
Service: Xuat NFI/RFI ra file Excel theo template chuan cua du an.

Template chuan (vd: RFI_NEW_20260527.xlsx) gom 2 sheet:
  - "RFI"         : trang bia (RFI No., Member Type, Inspector, Date...)
  - "MEMBER LIST" : danh sach cau kien can kiem tra
"""
from __future__ import annotations

import io
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

from streamlit_qc.core.db import DB


# Constants
RFI_CELL_NO = "C7"
RFI_CELL_MEMBER_TYPE = "C22"
RFI_CELL_INSPECTOR = "G59"
RFI_CELL_DATE = "F64"
RFI_CELL_ITP_ITEM = "H14"
RFI_CELL_ITP_DOC = "C19"

# Discipline checkbox cells (label nam ben phai, ghi "X" vao cell ben trai)
DISCIPLINE_CELLS = {
    "Mock-Up": "C11",
    "Raw Material": "J11",
    "Welding": "T11",
    "NDT": "AA11",
    "Dimension": "AE11",
    "Pre-assembly": "AM11",
    "Coating": "AV11",
}
DISCIPLINES = list(DISCIPLINE_CELLS.keys())

ML_SHEET = "MEMBER LIST"
ML_CELL_RFI_NO = "C7"
ML_DATA_START_ROW = 11

ML_COL_NO = 2
ML_COL_ITEM_NO = 3
ML_COL_DRAWING = 4
ML_COL_REV = 5
ML_COL_QTY = 6
ML_COL_WEIGHT = 7
ML_COL_UNIT = 8
ML_COL_STAGE = 9
ML_COL_LOCATION = 10
ML_COL_DDC_INS = 11
ML_COL_VERIFIED_BY = 12
ML_COL_ACC_REJ = 13
ML_COL_REMARK = 14
ML_COL_MILESTONE = 15


def _templates_dir() -> Path:
    base = Path(__file__).resolve().parent.parent / "data" / "templates"
    base.mkdir(parents=True, exist_ok=True)
    return base


def get_template_path(pid: int) -> Path:
    return _templates_dir() / f"project_{pid}" / "rfi_template.xlsx"


def _exports_dir(pid: int) -> Path:
    """Folder luu file NFI da xuat cua 1 du an."""
    base = Path(__file__).resolve().parent.parent / "data" / "exports" / f"project_{pid}"
    base.mkdir(parents=True, exist_ok=True)
    return base


def list_exported_nfis(pid: int) -> list[dict]:
    """Liet ke cac file NFI da xuat. Sort moi nhat len dau."""
    folder = _exports_dir(pid)
    out = []
    for f in folder.glob("*.xlsx"):
        try:
            stat = f.stat()
            out.append({
                "rfi_no": f.stem,
                "path": str(f),
                "size_bytes": stat.st_size,
                "ts": datetime.fromtimestamp(stat.st_mtime),
            })
        except OSError:
            continue
    out.sort(key=lambda x: x["ts"], reverse=True)
    return out


def read_exported_nfi(pid: int, rfi_no: str):
    """Doc lai file NFI da xuat theo rfi_no. None neu khong ton tai."""
    safe = re.sub(r"[\\/*?:\"<>|]", "_", rfi_no)
    path = _exports_dir(pid) / f"{safe}.xlsx"
    if path.exists():
        return path.read_bytes()
    return None


def delete_exported_nfi(pid: int, rfi_no: str) -> bool:
    """Xoa file NFI da xuat."""
    safe = re.sub(r"[\\/*?:\"<>|]", "_", rfi_no)
    path = _exports_dir(pid) / f"{safe}.xlsx"
    if path.exists():
        path.unlink()
        return True
    return False


def has_template(pid: int) -> bool:
    return get_template_path(pid).exists()


def save_template(pid: int, file_bytes: bytes) -> Path:
    path = get_template_path(pid)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(file_bytes)
    return path


def delete_template(pid: int) -> bool:
    path = get_template_path(pid)
    if path.exists():
        path.unlink()
        return True
    return False


def _extract_prefix_and_counter(rfi_no_text: str):
    s = str(rfi_no_text or "").strip()
    m = re.match(r"^(.+?)(\d+)\s*$", s)
    if m:
        return m.group(1), int(m.group(2))
    return s + "-", 0


def _peek_template_no(pid: int) -> str:
    path = get_template_path(pid)
    if not path.exists():
        return ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if "RFI" not in wb.sheetnames:
            return ""
        v = wb["RFI"][RFI_CELL_NO].value
        wb.close()
        return str(v or "")
    except Exception:
        return ""


def get_template_rfi_seed(pid: int):
    raw = _peek_template_no(pid)
    if not raw:
        return ("RFI-", 0)
    return _extract_prefix_and_counter(raw)


def _detect_pad_width(rfi_no_text: str) -> int:
    s = str(rfi_no_text or "").strip()
    m = re.search(r"(\d+)\s*$", s)
    if m:
        return len(m.group(1))
    return 3


def get_next_rfi_no_by_template(db, pid: int) -> str:
    raw_template_no = _peek_template_no(pid)
    prefix, template_counter = _extract_prefix_and_counter(raw_template_no)
    if not prefix:
        prefix = "RFI-"
    pad = _detect_pad_width(raw_template_no) or 3
    db_max = 0
    rows = db.conn.execute(
        "SELECT rfi_no FROM rfis WHERE project_id=? AND rfi_no LIKE ?",
        (pid, f"{prefix}%"),
    ).fetchall()
    for r in rows:
        _, n = _extract_prefix_and_counter(r["rfi_no"])
        if n > db_max:
            db_max = n
    next_n = max(template_counter, db_max) + 1
    return f"{prefix}{next_n:0{pad}d}"


def _fetch_components_for_export(db, pid: int, component_ids):
    if not component_ids:
        return []
    placeholders = ",".join("?" * len(component_ids))
    rows = db.conn.execute(
        f"SELECT id, code, data_json FROM components "
        f"WHERE project_id=? AND id IN ({placeholders})",
        [pid] + list(component_ids),
    ).fetchall()
    by_id = {r["id"]: r for r in rows}
    out = []
    for cid in component_ids:
        r = by_id.get(cid)
        if not r:
            continue
        try:
            d = json.loads(r["data_json"] or "{}")
        except (json.JSONDecodeError, TypeError):
            d = {}
        drawing = (
            d.get("manual_drawing") or d.get("drawing")
            or d.get("member_no") or d.get("section") or ""
        )
        mtype = d.get("type") or d.get("member_type") or ""
        if not mtype and r["code"]:
            m = re.search(r"[A-Z]{3,4}", r["code"])
            if m:
                mtype = m.group(0)
        out.append({
            "id": r["id"],
            "code": r["code"],
            "drawing": str(drawing),
            "rev": str(d.get("rev_no") or d.get("revision") or ""),
            "qty": d.get("qty") or 1,
            "weight": d.get("weight_kg") or d.get("weight") or "",
            "workshop": str(d.get("workshop") or ""),
            "member_type": str(mtype),
            "milestone": str(d.get("milestone") or ""),
        })
    return out


def export_rfi_file(
    db,
    pid,
    project_code,
    component_ids,
    user_name,
    inspection_stage="Fit-Up",
    persist_rfi_records=True,
    disciplines=None,
    itp_doc=None,
    itp_item_no=None,
    member_type_override=None,
    proposed_date=None,
):
    template_path = get_template_path(pid)
    if not template_path.exists():
        raise FileNotFoundError(
            f"Chua co template cho du an (pid={pid}). "
            f"Hay upload file RFI mau truoc khi xuat."
        )
    if not component_ids:
        raise ValueError("Chua chon cau kien nao de xuat NFI.")

    components = _fetch_components_for_export(db, pid, component_ids)
    if not components:
        raise ValueError("Khong tim thay cau kien nao trong DB voi ID da cho.")

    rfi_no = get_next_rfi_no_by_template(db, pid)
    wb = openpyxl.load_workbook(template_path)

    # Fill sheet RFI (trang bia)
    if "RFI" in wb.sheetnames:
        ws_rfi = wb["RFI"]
        ws_rfi[RFI_CELL_NO] = rfi_no
        # Member Type
        if member_type_override:
            ws_rfi[RFI_CELL_MEMBER_TYPE] = member_type_override
        else:
            mtypes = [c["member_type"] for c in components if c["member_type"]]
            if mtypes:
                ws_rfi[RFI_CELL_MEMBER_TYPE] = Counter(mtypes).most_common(1)[0][0]
        # Inspector
        if user_name:
            ws_rfi[RFI_CELL_INSPECTOR] = user_name
        # Date
        if proposed_date:
            if isinstance(proposed_date, str):
                try:
                    ws_rfi[RFI_CELL_DATE] = datetime.fromisoformat(proposed_date[:10])
                except (ValueError, TypeError):
                    ws_rfi[RFI_CELL_DATE] = proposed_date
            else:
                ws_rfi[RFI_CELL_DATE] = proposed_date
        else:
            ws_rfi[RFI_CELL_DATE] = datetime.now()
        # ITP Doc
        if itp_doc:
            ws_rfi[RFI_CELL_ITP_DOC] = str(itp_doc)
        # ITP Item no
        if itp_item_no not in (None, ""):
            try:
                ws_rfi[RFI_CELL_ITP_ITEM] = float(itp_item_no)
            except (ValueError, TypeError):
                ws_rfi[RFI_CELL_ITP_ITEM] = itp_item_no
        # Discipline checkboxes
        if disciplines:
            for disc in disciplines:
                cell = DISCIPLINE_CELLS.get(disc)
                if cell:
                    ws_rfi[cell] = "X"

    # Fill sheet MEMBER LIST
    if ML_SHEET not in wb.sheetnames:
        raise ValueError(f"Template thieu sheet '{ML_SHEET}'.")
    ws_ml = wb[ML_SHEET]
    ws_ml[ML_CELL_RFI_NO] = rfi_no

    DATA_START = ML_DATA_START_ROW
    last_data_row = DATA_START - 1
    for r in range(DATA_START, ws_ml.max_row + 1):
        v = ws_ml.cell(r, ML_COL_NO).value
        if isinstance(v, (int, float)):
            last_data_row = r
        else:
            if last_data_row >= DATA_START:
                break
    available_rows = max(0, last_data_row - DATA_START + 1)
    needed = len(components)

    if needed > available_rows:
        extra = needed - available_rows
        insert_at = max(last_data_row + 1, DATA_START)
        ws_ml.insert_rows(insert_at, amount=extra)

    for r in range(DATA_START, DATA_START + max(needed, available_rows)):
        for col in range(ML_COL_NO, ML_COL_MILESTONE + 1):
            ws_ml.cell(r, col).value = None

    for i, comp in enumerate(components):
        r = DATA_START + i
        ws_ml.cell(r, ML_COL_NO).value = i + 1
        ws_ml.cell(r, ML_COL_ITEM_NO).value = comp["code"]
        ws_ml.cell(r, ML_COL_DRAWING).value = comp["drawing"]
        ws_ml.cell(r, ML_COL_REV).value = comp["rev"]
        try:
            ws_ml.cell(r, ML_COL_QTY).value = int(comp["qty"]) if comp["qty"] else 1
        except (ValueError, TypeError):
            ws_ml.cell(r, ML_COL_QTY).value = comp["qty"] or 1
        try:
            ws_ml.cell(r, ML_COL_WEIGHT).value = (
                float(comp["weight"]) if comp["weight"] else None
            )
        except (ValueError, TypeError):
            ws_ml.cell(r, ML_COL_WEIGHT).value = comp["weight"] or None
        ws_ml.cell(r, ML_COL_UNIT).value = "PC"
        ws_ml.cell(r, ML_COL_STAGE).value = inspection_stage
        ws_ml.cell(r, ML_COL_LOCATION).value = comp["workshop"]
        ws_ml.cell(r, ML_COL_DDC_INS).value = user_name or ""
        ws_ml.cell(r, ML_COL_MILESTONE).value = comp["milestone"]

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    file_bytes = buf.getvalue()

    # Luu file vao exports/ de xem lai sau
    try:
        safe_no = re.sub(r"[\\/*?:\"<>|]", "_", rfi_no)
        out_path = _exports_dir(pid) / f"{safe_no}.xlsx"
        out_path.write_bytes(file_bytes)
    except Exception as e:
        print(f"[rfi_export] save to exports failed: {e}")

    if persist_rfi_records:
        try:
            today_iso = datetime.now().date().isoformat()
            inspection_type = (
                "FUR" if inspection_stage.lower().startswith("fit") else "DGRP"
            )
            codes_csv = ",".join(c["code"] for c in components)
            note_attached = f"Exported {len(components)} ck: {codes_csv[:480]}"
            db.conn.execute(
                "INSERT INTO rfis "
                "(project_id, component_id, rfi_no, inspection_type, "
                " proposed_date, submitted_by, response_note, status) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, 'SUBMITTED') "
                "ON CONFLICT DO NOTHING",
                (pid, components[0]["id"], rfi_no, inspection_type,
                 today_iso, user_name or "", note_attached),
            )
            db.conn.commit()
        except Exception as e:
            print(f"[rfi_export] insert rfis failed: {e}")
        try:
            db.log(
                user_name or "",
                "RFI_EXPORT",
                "rfis",
                None,
                f"rfi_no={rfi_no} n={len(components)} stage={inspection_stage}",
            )
        except Exception as e:
            print(f"[rfi_export] log failed: {e}")

    return file_bytes, rfi_no
