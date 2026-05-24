# -*- coding: utf-8 -*-
"""
Service xuất Excel THƯ MỜI NGHIỆM THU.

Output Excel gồm 2 sheet:
  - Sheet "Thư mời": header (Kính gửi, dự án, ngày, công đoạn, số RFI)
  - Sheet "Danh sách cấu kiện": bảng danh sách cấu kiện kèm thông tin
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from streamlit_qc.core.db import DB


STAGE_LABELS = {
    "FUR": "Fit-up Report (Nghiệm thu gá lắp)",
    "DIR": "Dimension Report (Nghiệm thu kích thước)",
    "VIR": "Visual Inspection (Nghiệm thu mối hàn ngoại quan)",
    "NDT": "NDT (Kiểm tra không phá huỷ: MT/PT/UT/RT)",
    "PAINT": "Paint Inspection (Nghiệm thu sơn)",
    "DGRP": "Đóng gói / Xuất xưởng",
}


@dataclass
class InvitationData:
    project_code: str = ""
    project_name: str = ""
    stage: str = "FUR"
    inspection_date: str = ""
    location: str = "Nhà máy Đại Dũng — Long An"
    rfi_no: str = ""
    recipient: str = "Kính gửi: Quý giám sát"
    sender: str = "Phòng QC — Đại Dũng Steel"
    note: str = ""
    components: list[dict] = field(default_factory=list)


# ====================================================================
# DATA QUERY
# ====================================================================
def get_components_for_stage(
    db: DB,
    pid: int,
    component_ids: list[int] | None = None,
    status_filter: list[str] | None = None,
    limit: int = 1000,
) -> list[dict]:
    """Lấy danh sách cấu kiện kèm thông tin chi tiết để đưa vào thư mời."""
    if component_ids:
        placeholders = ",".join("?" * len(component_ids))
        q = (
            f"SELECT id, code, data_json, status FROM components "
            f"WHERE project_id=? AND id IN ({placeholders}) "
            f"ORDER BY code LIMIT ?"
        )
        params = [pid, *component_ids, limit]
    elif status_filter:
        placeholders = ",".join("?" * len(status_filter))
        q = (
            f"SELECT id, code, data_json, status FROM components "
            f"WHERE project_id=? AND status IN ({placeholders}) "
            f"ORDER BY code LIMIT ?"
        )
        params = [pid, *status_filter, limit]
    else:
        q = (
            "SELECT id, code, data_json, status FROM components "
            "WHERE project_id=? ORDER BY code LIMIT ?"
        )
        params = [pid, limit]

    rows = db.conn.execute(q, params).fetchall()
    out = []
    for r in rows:
        try:
            d = json.loads(r["data_json"]) if r["data_json"] else {}
        except (json.JSONDecodeError, TypeError):
            d = {}
        out.append({
            "id": r["id"],
            "code": r["code"],
            "status": r["status"],
            "drawing": d.get("manual_drawing") or d.get("drawing") or "",
            "workshop": d.get("workshop") or "",
            "weight_kg": d.get("weight_kg") or "",
            "surface_m2": d.get("surface_m2") or "",
            "qty": d.get("qty") or 1,
            "material": d.get("material_code") or "",
            "note": d.get("note") or "",
        })
    return out


# ====================================================================
# EXCEL EXPORT
# ====================================================================
def build_invitation_excel(inv: InvitationData) -> bytes:
    """Tạo file Excel thư mời, trả về bytes."""
    wb = Workbook()

    # ===== Sheet 1: Thư mời =====
    ws = wb.active
    ws.title = "Thư mời"

    NAVY = "0F1E40"
    GOLD = "D4A744"
    GRAY_LIGHT = "F2F2F2"

    thin = Side(style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    bold_navy = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, color="FFFFFF", size=16)
    label_font = Font(bold=True, color=NAVY, size=11)
    value_font = Font(color="333333", size=11)
    italic_gold = Font(italic=True, color="8B6914", size=10)

    fill_navy = PatternFill("solid", fgColor=NAVY)
    fill_gold = PatternFill("solid", fgColor=GOLD)
    fill_gray = PatternFill("solid", fgColor=GRAY_LIGHT)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title banner
    ws.merge_cells("A1:F2")
    cell = ws["A1"]
    cell.value = "THƯ MỜI NGHIỆM THU CẤU KIỆN"
    cell.font = title_font
    cell.fill = fill_navy
    cell.alignment = center

    ws.merge_cells("A3:F3")
    sub = ws["A3"]
    sub.value = "ĐẠI DŨNG STEEL — PHÒNG QC"
    sub.font = italic_gold
    sub.alignment = center

    # Info table
    rows = [
        ("Dự án:", f"{inv.project_code} — {inv.project_name}"),
        ("Công đoạn nghiệm thu:", STAGE_LABELS.get(inv.stage, inv.stage)),
        ("Ngày dự kiến:", inv.inspection_date or ""),
        ("Địa điểm:", inv.location or ""),
        ("Số RFI:", inv.rfi_no or ""),
        ("Kính gửi:", inv.recipient or ""),
        ("Đơn vị mời:", inv.sender or ""),
    ]
    start = 5
    for i, (lbl, val) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=lbl).font = label_font
        ws.cell(row=r, column=1).fill = fill_gray
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = left
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=val)
        c.font = value_font
        c.alignment = left
        c.border = border
        for j in range(3, 7):
            ws.cell(row=r, column=j).border = border

    # Body letter
    body_start = start + len(rows) + 2
    ws.merge_cells(start_row=body_start, start_column=1, end_row=body_start, end_column=6)
    body = ws.cell(
        row=body_start, column=1,
        value=(
            "Phòng QC – Đại Dũng Steel trân trọng kính mời Quý Giám sát đến nghiệm thu "
            f"công đoạn nêu trên cho {len(inv.components)} cấu kiện trong bảng đính kèm "
            "(xem sheet \"Danh sách cấu kiện\")."
        ),
    )
    body.font = value_font
    body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[body_start].height = 50

    if inv.note:
        note_row = body_start + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
        nc = ws.cell(row=note_row, column=1, value=f"Ghi chú: {inv.note}")
        nc.font = italic_gold
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Signature block
    sig_row = body_start + 5
    ws.cell(row=sig_row, column=1, value="Đại diện Phòng QC").font = label_font
    ws.cell(row=sig_row, column=4, value="Đại diện Giám sát").font = label_font
    ws.cell(row=sig_row, column=1).alignment = center
    ws.cell(row=sig_row, column=4).alignment = center
    ws.cell(row=sig_row + 1, column=1, value="(Ký, ghi rõ họ tên)").font = italic_gold
    ws.cell(row=sig_row + 1, column=4, value="(Ký, ghi rõ họ tên)").font = italic_gold
    ws.cell(row=sig_row + 1, column=1).alignment = center
    ws.cell(row=sig_row + 1, column=4).alignment = center

    # Column widths
    widths = [22, 18, 18, 18, 18, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ===== Sheet 2: Danh sách cấu kiện =====
    ws2 = wb.create_sheet("Danh sách cấu kiện")
    headers = [
        "STT", "Mã cấu kiện", "Bản vẽ", "Phân xưởng",
        "Vật liệu", "Khối lượng (kg)", "Diện tích (m²)",
        "Số lượng", "Trạng thái", "Ghi chú",
    ]
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = bold_navy
        c.fill = fill_navy
        c.alignment = center
        c.border = border

    for i, comp in enumerate(inv.components, 1):
        row = i + 1
        vals = [
            i,
            comp.get("code", ""),
            comp.get("drawing", ""),
            comp.get("workshop", ""),
            comp.get("material", ""),
            comp.get("weight_kg", ""),
            comp.get("surface_m2", ""),
            comp.get("qty", 1),
            comp.get("status", ""),
            comp.get("note", ""),
        ]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=j, value=v)
            c.alignment = left if j in (2, 3, 4, 5, 10) else center
            c.border = border
            c.font = value_font
            if i % 2 == 0:
                c.fill = fill_gray

    # Column widths sheet 2
    widths2 = [6, 22, 18, 12, 14, 14, 14, 10, 14, 30]
    for i, w in enumerate(widths2):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    ws2.freeze_panes = "A2"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def default_filename(inv: InvitationData) -> str:
    """Sinh tên file gợi ý."""
    today = date.today().isoformat()
    pcode = (inv.project_code or "PROJ").replace(" ", "_")
    stage = inv.stage or "STAGE"
    return f"ThuMoiNghiemThu_{pcode}_{stage}_{today}.xlsx"
