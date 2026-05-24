# -*- coding: utf-8 -*-
"""
Service: NCR Management (Non-Conformance Report).

Workflow chuẩn:
  OPEN → IN_REVIEW → RESOLVED → CLOSED
  (CLOSED là trạng thái cuối, không quay lại)

Mỗi NCR gắn với 1 dự án và (tuỳ chọn) 1 cấu kiện cụ thể.
Severity: LOW / MEDIUM / HIGH / CRITICAL.
"""
from __future__ import annotations

import io
from datetime import datetime, date
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from streamlit_qc.core.db import DB


# ====================================================================
# CONSTANTS
# ====================================================================
NCR_STATUSES = ("OPEN", "IN_REVIEW", "RESOLVED", "CLOSED")
NCR_SEVERITIES = ("LOW", "MEDIUM", "HIGH", "CRITICAL")

STATUS_LABEL = {
    "OPEN":      "🔴 Mở (Open)",
    "IN_REVIEW": "🟡 Đang xem xét",
    "RESOLVED":  "🟢 Đã xử lý",
    "CLOSED":    "⚫ Đóng",
}

SEVERITY_LABEL = {
    "LOW":      "🟢 Thấp",
    "MEDIUM":   "🟡 Trung bình",
    "HIGH":     "🟠 Cao",
    "CRITICAL": "🔴 Nghiêm trọng",
}

STATUS_COLOR = {
    "OPEN":      "DC2626",   # red
    "IN_REVIEW": "D97706",   # amber
    "RESOLVED":  "16A34A",   # green
    "CLOSED":    "6B7280",   # gray
}


# ====================================================================
# CRUD
# ====================================================================
def create_ncr(
    db: DB,
    pid: int,
    title: str,
    description: str = "",
    component_code: str | None = None,
    severity: str = "MEDIUM",
    deadline: str | None = None,
    raised_by: str | None = None,
    ncr_no: str | None = None,
) -> tuple[int, str]:
    """Tạo NCR mới. Tự sinh ncr_no nếu không truyền. Trả về (id, ncr_no)."""
    component_id = None
    if component_code:
        row = db.find_component(pid, str(component_code).strip())
        if row:
            component_id = row["id"]

    if not ncr_no:
        ncr_no = db.get_next_ncr_no(pid)

    nid = db.add_ncr(
        pid=pid,
        ncr_no=ncr_no,
        title=title,
        description=description,
        component_id=component_id,
        severity=severity,
        deadline=deadline,
        raised_by=raised_by,
    )
    db.conn.commit()
    return nid, ncr_no


def list_ncrs_df(
    db: DB,
    pid: int,
    status: str | None = None,
    severity: str | None = None,
) -> pd.DataFrame:
    """Trả về DataFrame các NCR cho UI."""
    rows = db.list_ncrs(pid, status=status, severity=severity)
    if not rows:
        return pd.DataFrame()
    out = []
    for r in rows:
        out.append({
            "ID": r["id"],
            "Số NCR": r["ncr_no"],
            "Tiêu đề": r["title"],
            "Cấu kiện": r["component_code"] or "—",
            "Mức độ": SEVERITY_LABEL.get(r["severity"], r["severity"]),
            "Trạng thái": STATUS_LABEL.get(r["status"], r["status"]),
            "Người báo": r["raised_by"] or "",
            "Deadline": r["deadline"] or "",
            "Người xử lý": r["resolved_by"] or "",
            "Ngày xử lý": str(r["resolved_at"] or "")[:10],
            "Mô tả": r["description"] or "",
            "Nguyên nhân": r["root_cause"] or "",
            "Hành động sửa": r["corrective_action"] or "",
            "Ngày tạo": str(r["created_at"])[:16].replace("T", " "),
        })
    return pd.DataFrame(out)


def update_status(
    db: DB,
    ncr_id: int,
    new_status: str,
    resolved_by: str = "",
    root_cause: str = "",
    corrective_action: str = "",
) -> None:
    """Đổi trạng thái NCR + ghi root cause / corrective action nếu cần."""
    db.update_ncr_status(
        ncr_id=ncr_id,
        new_status=new_status,
        resolved_by=resolved_by or None,
        root_cause=root_cause or None,
        corrective_action=corrective_action or None,
    )
    db.conn.commit()


def delete(db: DB, ncr_id: int) -> None:
    db.delete_ncr(ncr_id)
    db.conn.commit()


def counts(db: DB, pid: int) -> dict[str, int]:
    return db.count_ncrs_by_status(pid)


# ====================================================================
# EXCEL EXPORT
# ====================================================================
def export_to_excel(db: DB, pid: int, project_code: str = "") -> bytes:
    """Xuất Excel danh sách NCR — đẹp + có conditional fill theo status."""
    df = list_ncrs_df(db, pid)
    wb = Workbook()
    ws = wb.active
    ws.title = "NCR List"

    NAVY = "0F1E40"
    thin = Side(style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Header
    if df.empty:
        ws.cell(row=1, column=1, value="Không có NCR nào.")
    else:
        headers = list(df.columns)
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for i, row in enumerate(df.itertuples(index=False), 2):
            for j, v in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = border
                c.font = Font(size=10)
            # Color row by status (column 6 = "Trạng thái")
            status_label = str(row[5]) if len(row) > 5 else ""
            for st_key, st_color in STATUS_COLOR.items():
                if STATUS_LABEL[st_key] in status_label:
                    for j in range(1, len(headers) + 1):
                        ws.cell(row=i, column=j).fill = PatternFill(
                            "solid", fgColor=f"{st_color}22"  # light tint
                        )
                    break
        # Column widths
        widths = {1:6, 2:14, 3:30, 4:14, 5:14, 6:16, 7:14, 8:12, 9:14, 10:12, 11:30, 12:30, 13:30, 14:18}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
