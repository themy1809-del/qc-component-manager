# -*- coding: utf-8 -*-
"""
Service: Batch Handover — gom lô cấu kiện ACCEPTED + QR code + Packing list.

State machine:
  DRAFT → READY → DELIVERED → CONFIRMED
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from streamlit_qc.core.db import DB


BATCH_STATUSES = ("DRAFT", "READY", "DELIVERED", "CONFIRMED")
STATUS_LABEL = {
    "DRAFT":     "📝 Nháp",
    "READY":     "📦 Sẵn sàng",
    "DELIVERED": "🚛 Đã giao",
    "CONFIRMED": "✅ Xác nhận",
}


def create_batch_from_components(
    db: DB,
    pid: int,
    project_code: str,
    component_ids: list[int],
    created_by: str,
    notes: str = "",
    allow_not_accepted: bool = False,
) -> tuple[int, str, int]:
    """
    Tạo batch mới từ danh sách CK. Trả về (batch_id, batch_no, n_items).

    Mặc định CHỈ cho phép gom CK đang ACCEPTED — set allow_not_accepted=True
    nếu muốn gom cả CK chưa ACCEPTED (hiếm dùng).
    """
    if not component_ids:
        raise ValueError("Cần ít nhất 1 cấu kiện.")

    # Validate
    placeholders = ",".join("?" * len(component_ids))
    rows = db.conn.execute(
        f"SELECT id, code, status, data_json FROM components "
        f"WHERE project_id=? AND id IN ({placeholders})",
        (pid, *component_ids),
    ).fetchall()

    if not allow_not_accepted:
        not_ok = [r for r in rows if r["status"] != "ACCEPTED"]
        if not_ok:
            raise ValueError(
                f"{len(not_ok)} cấu kiện chưa ACCEPTED: "
                + ", ".join(r["code"] for r in not_ok[:5])
                + ("..." if len(not_ok) > 5 else "")
            )

    # Tổng khối lượng
    total_w = 0.0
    for r in rows:
        try:
            d = json.loads(r["data_json"])
            w = float(d.get("weight_kg") or 0)
            total_w += w
        except (json.JSONDecodeError, TypeError, ValueError):
            continue

    batch_no = db.get_next_batch_no(pid, project_code)
    batch_id = db.add_batch(pid, batch_no, total_weight_kg=total_w,
                            notes=notes, created_by=created_by)
    n_items = db.add_batch_items(batch_id, [r["id"] for r in rows])
    db.conn.commit()
    db.log(created_by, "BATCH_CREATE", "batches", batch_id,
           f"batch_no={batch_no} n={n_items} w={total_w:.1f}kg")
    return batch_id, batch_no, n_items


def list_batches_df(db: DB, pid: int, status: str | None = None) -> pd.DataFrame:
    rows = db.list_batches(pid, status=status)
    if not rows:
        return pd.DataFrame()
    out = []
    for r in rows:
        # Count items
        cnt_row = db.conn.execute(
            "SELECT COUNT(*) c FROM batch_items WHERE batch_id=?", (r["id"],),
        ).fetchone()
        out.append({
            "ID": r["id"],
            "Số lô": r["batch_no"],
            "Trạng thái": STATUS_LABEL.get(r["status"], r["status"]),
            "Số CK": cnt_row["c"] if cnt_row else 0,
            "Khối lượng (kg)": float(r["total_weight_kg"] or 0),
            "Ngày bàn giao": r["handover_date"] or "",
            "Người nhận": r["receiver_name"] or "",
            "Công ty nhận": r["receiver_company"] or "",
            "Người tạo": r["created_by"] or "",
            "Ngày tạo": str(r["created_at"])[:16].replace("T", " "),
            "Ghi chú": r["notes"] or "",
        })
    return pd.DataFrame(out)


def transition_status(db: DB, batch_id: int, new_status: str,
                      by_user: str = "",
                      handover_date: str | None = None,
                      receiver_name: str | None = None,
                      receiver_company: str | None = None) -> None:
    db.update_batch_status(
        batch_id, new_status,
        handover_date=handover_date,
        receiver_name=receiver_name,
        receiver_company=receiver_company,
    )
    db.conn.commit()
    db.log(by_user, f"BATCH_{new_status.upper()}", "batches", batch_id, "")


# ====================================================================
# QR CODE
# ====================================================================
def gen_qr_for_component(component_id: int, component_code: str,
                          app_url: str = "https://qc-daidung.streamlit.app") -> bytes:
    """Tạo QR code PNG bytes cho 1 cấu kiện. Fallback nếu thiếu qrcode."""
    try:
        import qrcode
    except ImportError:
        return b""
    payload = json.dumps({
        "id": component_id,
        "code": component_code,
        "app": app_url,
    }, ensure_ascii=False)
    qr = qrcode.QRCode(version=1, box_size=6, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


# ====================================================================
# PACKING LIST (Excel)
# ====================================================================
def build_packing_list_excel(db: DB, batch_id: int) -> bytes:
    """Tạo Excel packing list cho 1 batch."""
    batch_row = db.conn.execute(
        "SELECT b.*, p.code AS project_code, p.name AS project_name "
        "FROM batches b JOIN projects p ON p.id = b.project_id "
        "WHERE b.id=?", (batch_id,),
    ).fetchone()
    if not batch_row:
        raise ValueError(f"Batch ID {batch_id} không tồn tại")

    items = db.list_batch_items(batch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"

    NAVY = "0F1E40"
    GOLD = "D4A744"
    thin = Side(style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "PACKING LIST — DAI DUNG STEEL"
    c.font = Font(bold=True, color="FFFFFF", size=18)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Info
    info = [
        ("Số lô / Batch No:",       batch_row["batch_no"]),
        ("Dự án / Project:",        f"{batch_row['project_code']} — {batch_row['project_name']}"),
        ("Trạng thái / Status:",    batch_row["status"]),
        ("Ngày bàn giao / Date:",   batch_row["handover_date"] or "—"),
        ("Người nhận / Receiver:",  batch_row["receiver_name"] or "—"),
        ("Công ty / Company:",      batch_row["receiver_company"] or "—"),
        ("Tổng khối lượng / Total:", f"{float(batch_row['total_weight_kg'] or 0):,.2f} kg"),
        ("Tổng số CK / Items:",     str(len(items))),
    ]
    for i, (lbl, val) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True, color=NAVY, size=11)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        c = ws.cell(row=i, column=2, value=val)
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Items table
    table_start = 3 + len(info) + 2
    headers = ["STT", "Mã cấu kiện", "Bản vẽ", "Xưởng", "Khối lượng (kg)",
               "Số lượng", "Trạng thái", "Ghi chú"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=table_start, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, item in enumerate(items, 1):
        try:
            d = json.loads(item["data_json"])
        except (json.JSONDecodeError, TypeError):
            d = {}
        row = table_start + i
        vals = [
            i,
            item["component_code"],
            d.get("manual_drawing") or d.get("drawing") or "",
            d.get("workshop") or "",
            float(d.get("weight_kg") or 0),
            item["quantity"],
            item["status"],
            "",
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j in (1, 6) else "left",
                                    vertical="center", wrap_text=True)
            c.border = border
            c.font = Font(size=10)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F8FAFC")

    # Signature
    sig_row = table_start + len(items) + 3
    ws.cell(row=sig_row, column=1, value="BÊN GIAO (QC Đại Dũng)").font = Font(bold=True, color=NAVY)
    ws.cell(row=sig_row, column=5, value="BÊN NHẬN").font = Font(bold=True, color=NAVY)
    ws.cell(row=sig_row + 3, column=1, value="_______________________")
    ws.cell(row=sig_row + 3, column=5, value="_______________________")
    ws.cell(row=sig_row + 4, column=1,
            value=f"Date: {date.today().strftime('%d/%m/%Y')}").font = Font(italic=True, color="666666")
    ws.cell(row=sig_row + 4, column=5, value="Date: ____/____/______").font = Font(italic=True, color="666666")

    # Column widths
    widths = [6, 22, 18, 12, 16, 10, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{table_start + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
