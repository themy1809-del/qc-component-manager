# -*- coding: utf-8 -*-
"""
Engine đọc Excel (xlsb/xlsx/xlsm/xls/csv) + auto-detect header/cột thông minh.

Copy 100% logic từ Tkinter v1.0.2 (dòng 76-216), chỉ refactor thành hàm pure.
KHÔNG đổi SMART_KEYWORDS dictionary (xem core/constants.py).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from streamlit_qc.core.constants import SMART_KEYWORDS


def read_excel_any(
    path: str | Path,
    sheet_name: str | int | None = None,
    header: int = 0,
) -> pd.DataFrame:
    """
    Đọc 1 sheet Excel/CSV. Hỗ trợ .xlsb, .xlsx, .xlsm, .xls, .csv.

    Args:
        path: Đường dẫn file.
        sheet_name: Tên sheet hoặc index, None = sheet đầu.
        header: Dòng tiêu đề (0-based).

    Returns:
        DataFrame.

    Raises:
        ValueError: Nếu phần mở rộng không hỗ trợ.
    """
    ext = Path(path).suffix.lower()
    if ext == ".xlsb":
        return pd.read_excel(path, sheet_name=sheet_name, header=header, engine="pyxlsb")
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(path, sheet_name=sheet_name, header=header, engine="openpyxl")
    if ext == ".xls":
        return pd.read_excel(path, sheet_name=sheet_name, header=header)
    if ext == ".csv":
        return pd.read_csv(path, header=header)
    raise ValueError(f"Định dạng không hỗ trợ: {ext}")


def list_sheet_names(path: str | Path, visible_only: bool = True) -> list[str]:
    """
    Liệt kê tên các sheet trong workbook Excel.

    Args:
        path: Đường dẫn file.
        visible_only: Nếu True (mặc định), chỉ trả về sheet VISIBLE
            (bỏ qua hidden + veryHidden). File daily QC thường có nhiều
            sheet tham chiếu hidden — QC không cần thấy.

    Returns:
        List tên sheet.
    """
    ext = Path(path).suffix.lower()
    if ext == ".xlsb":
        # pyxlsb không phân biệt được hidden/visible → trả về tất cả
        from pyxlsb import open_workbook
        with open_workbook(path) as wb:
            return list(wb.sheets)
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if visible_only:
                # Chỉ lấy sheet có sheet_state == 'visible'
                # (loại bỏ 'hidden' và 'veryHidden')
                names = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
                # Fallback: nếu toàn bộ đều bị ẩn → trả về tất cả để khỏi rỗng
                if not names:
                    return list(wb.sheetnames)
                return names
            return list(wb.sheetnames)
        finally:
            wb.close()
    return ["Sheet1"]


def smart_detect_header_row(file_path: str | Path, sheet_name: str | int | None) -> int:
    """
    Tự dò dòng tiêu đề trong 20 dòng đầu.

    Logic chấm điểm:
      +5 mỗi keyword QC mạnh ("tên cấu kiện", "member no", "punch no", "stt", "kết quả", ...)
      +3 mỗi keyword SMART_KEYWORDS chung
      +1 mỗi ô là chuỗi text ngắn (2-60 ký tự, không phải số)
      -3 nếu dòng có >50% ô là số (dòng dữ liệu, không phải header)
      -2 nếu dòng chứa cụm typical metadata ("công trình", "hạng mục", "ngày", "rfi số")

    Returns:
        Index dòng tiêu đề (0-based). Mặc định 0 nếu không tìm được.
    """
    try:
        df_preview = read_excel_any(file_path, sheet_name=sheet_name, header=None)
        df_preview = df_preview.head(20)
    except Exception:
        return 0

    # Keyword mạnh — chỉ xuất hiện ở header thật
    STRONG_KEYWORDS = [
        "tên cấu kiện", "ten cau kien", "member no", "punch no", "member punch",
        "mã cấu kiện", "ma cau kien", "piece mark", "kết quả", "ket qua",
        "result", "drawing no", "stt",
    ]
    METADATA_NOISE = [
        "công trình", "cong trinh", "hạng mục", "hang muc", "danh sách",
        "địa điểm", "dia diem", "nội dung", "thời gian", "rfi số", "rfi so",
        "qc phụ trách", "thông tin liên hệ", "ngày:", "company", "project",
    ]

    keywords_all: list[str] = []
    for kws in SMART_KEYWORDS.values():
        keywords_all.extend(kws)

    best_row, best_score = 0, -1

    for i, row in df_preview.iterrows():
        cells = [
            str(v).strip().lower()
            for v in row
            if v is not None and str(v).strip() and not pd.isna(v)
        ]
        if len(cells) < 3:
            continue

        # Đếm số ô là pure number (data row)
        num_count = sum(
            1 for c in cells
            if c.replace(".", "").replace(",", "").replace("-", "").isdigit()
        )
        if num_count > len(cells) / 2:
            # Hơn 50% là số → đây là dòng dữ liệu, không phải header
            continue

        score = 0

        # +1 cho mỗi ô text ngắn
        for c in cells:
            if 2 <= len(c) <= 60 and not c.replace(".", "").replace(",", "").isdigit():
                score += 1

        # Strong keyword (header thật mới có)
        row_text = " | ".join(cells)
        for kw in STRONG_KEYWORDS:
            if kw in row_text:
                score += 5

        # Generic SMART keywords
        for c in cells:
            for kw in keywords_all:
                if kw in c:
                    score += 3
                    break

        # Trừ điểm nếu dòng có metadata noise (Công trình, Hạng mục, ...)
        for noise in METADATA_NOISE:
            if noise in row_text:
                score -= 2

        if score > best_score:
            best_score, best_row = score, i

    return best_row


def smart_match_columns(
    headers: list[str],
    fields: list[str],
) -> dict[str, str]:
    """
    Tự match header Excel → tên trường hệ thống dựa SMART_KEYWORDS.

    Khác bản Tkinter ở chỗ trả về dict thay vì set widget value
    (Streamlit không có widget object như tkinter).

    Args:
        headers: Danh sách tên cột thực tế từ file Excel.
        fields: Danh sách trường cần map (vd ['code', 'name', 'workshop', ...]).

    Returns:
        Dict {field: header_đã_match}. Không match được thì field không có key.
    """
    used_headers: set[str] = set()
    result: dict[str, str] = {}

    norm_headers = [
        (h, str(h).replace("\n", " ").replace("_", " ").strip().lower())
        for h in headers
    ]

    # Loại header có hậu tố "cũ/old/backup/mới/new" cho các field nhận diện chính
    # (vd: file Bison có "Tên cấu kiện cũ" — chỉ 46 unique values, là drawing parent
    # KHÔNG phải piece-level code). Tránh smart-match nhầm vào các cột này.
    EXCLUDE_SUFFIX = ("cu", "cũ", "old", "backup", "moi", "mới", "new", "version", "history")
    # Loại header bắt đầu bằng "kiểm tra"/"check" (cột kiểm tra ô check, không phải data)
    EXCLUDE_PREFIX = ("kiểm tra", "kiem tra", "check", "ktra")
    EXCLUDE_FIELDS = {"code", "name", "member_no", "drawing"}

    for field in fields:
        keywords = SMART_KEYWORDS.get(field)
        if not keywords:
            continue
        best_h, best_score = None, 0
        for h, norm in norm_headers:
            if h in used_headers:
                continue
            # Skip cột "...cũ"/"...old"/"...mới" hoặc bắt đầu "kiểm tra"/"check" cho field nhận diện chính
            if field in EXCLUDE_FIELDS:
                tokens = norm.split()
                if any(t in EXCLUDE_SUFFIX for t in tokens):
                    continue
                if any(norm.startswith(pre) for pre in EXCLUDE_PREFIX):
                    continue
            for rank, kw in enumerate(keywords):
                kw_norm = kw.lower()
                # Scoring: exact match được floor ≥ 90 (luôn thắng substring 80)
                if norm == kw_norm:
                    score = max(100 - rank * 2, 90)
                elif kw_norm in norm:
                    score = 80 - rank * 2
                elif norm in kw_norm and len(norm) >= 3:
                    score = 50 - rank * 2
                else:
                    continue
                if score > best_score:
                    best_score, best_h = score, h
        if best_h is not None:
            result[field] = best_h
            used_headers.add(best_h)

    return result
